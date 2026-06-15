import pandas as pd
from datetime import datetime,date
import os,re
from element import printElement
from informations.charge_records import records
from informations.clients_status import clients
from utils import convert_to_full_date
from copy import deepcopy
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from decimal import Decimal, getcontext, ROUND_DOWN
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#
summary_headers        = ['户名','查看人数','查看数','人均查看','点击数','单次点击','注册数','单次注册','充值数','单次充值','户ID','花销']
summary_concat_headers = ['剩余额度','总额度','现金充值','使用备用金','状态','已清算']
# '户名','查看人数','查看数','人均查看','点击数','单次点击','注册数','单次注册','充值数','单次充值','户ID','花销'
#  0       1       2         3         4       5        6       7         8      9         10    11
# '剩余额度','总额度','现金充值','储备金充值','状态','已清算'
#  12        13       14           15       16    17


everyday_headers        = ['户名','查看人数','查看数','人均查看','点击数','单次点击','注册数','单次注册','充值数','单次充值','户ID','花销','剩余额度','日期']
everyday_concat_headers = ['总充值']

adset_headers = ['系列名称','展示次数','链接点击量','链接点击率','完成注册次数','单次完成注册费用','购物次数','单次购物费用','开始试用次数','单次开始试用费用','已花费金额','日期','占位1','占位2']
# '系列名称','展示次数','链接点击量','链接点击率','完成注册次数','单次完成注册费用','购物次数','单次购物费用','开始试用次数','单次开始试用费用','已花费金额','广告系列编号'
#     0       1         2          3            4           5                6           7           8               9            10          11


yesterday_concat_headers = ['状态','已清算']
# '户名','查看人数','查看数','人均查看','点击数','单次点击','注册数','单次注册','充值数','单次充值','户ID','花销'
#  0       1       2         3         4       5        6       7         8      9         10    11
# '状态','已清算'
#  12      13


account_info_headers = {
    "name": "名称",
    "spending_limit": "每日限额",
    "threshold": "结算阈值",
    "status": "状态",
    "currency": "货币",
}

output_folder = "statistics"
os.makedirs(output_folder, exist_ok=True)

def extract_float(val):
    """Clean and convert any value to float if possible"""
    if pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).strip().replace(',', '')
    match = re.search(r'-?\d+\.?\d*', val)
    return float(match.group()) if match else None

def collectChargesByAccount(flat_records):
    charge_map = defaultdict(list)
    for r in flat_records:
        acc_id = r['id']
        charge_map[acc_id].append((r['amount'],  r['date'], '是' if r['deposit'] else '-'))
    return charge_map

def collectStatusByAccount():
    status_map = defaultdict(dict)
    for client in clients:
        acc_id = client['id']
        status_map[acc_id]['option'] = client['option']
        status_map[acc_id]['cleared'] = client['cleared']

    return status_map

def get_end_date():
    """Read end date from file, if not exist return today"""
    try:
        with open('end_date.txt', 'r') as f:
            end_date = f.read().strip()
            return datetime.strptime(end_date, '%Y-%m-%d').date()
    except FileNotFoundError:
        return datetime.today().date()

def writeReport(report, bm_ins):
    print(report)

    concated_report = concatSummaryTable(report,bm_ins['business_name'])
    headers = summary_headers + summary_concat_headers
    try:
        df = pd.DataFrame(data=concated_report, columns=headers)
    except:
        printElement(concated_report)
        df = None

    df['display'] = ''

    status_col = df.columns[16]
    cleared_col = df.columns[17]
    display_col = df.columns[18]

    df[display_col] = df[[status_col,cleared_col]].apply(lambda x: not (x[0] == True and x[1] == False), axis=1)

    grouped_rows = []

    for display, group in df.groupby(display_col):
        group = group.copy().reset_index(drop=True)

        # Convert values that can be numeric
        for col in group.columns:
            group[col] = pd.to_numeric(group[col], errors='ignore')

        total_col = group.columns[13]
        cash_col = group.columns[14]
        deposit_col = group.columns[15]
        spent_col = group.columns[11]
        remaining_col = group.columns[12]

        cash_col_d = pd.to_numeric(group[cash_col], errors='coerce').fillna(0)
        deposit_col_d = pd.to_numeric(group[deposit_col], errors='coerce').fillna(0)
        total_col_d = cash_col_d + deposit_col_d # 13
        group[total_col] = total_col_d # write to column 13

        spent_col_d = pd.to_numeric(group[spent_col], errors='coerce').fillna(0)
        remaining_col_d = round((total_col_d - spent_col_d), 2)  # 12

        group[remaining_col] = remaining_col_d

        status_col = group.columns[16]
        cleared_col = group.columns[17]

        # Step 1: Mark '正常' or '异常' based on original column 14 values
        group[status_col] = group[status_col].apply(lambda x: '正常' if x else '异常')
        group[cleared_col] = group[cleared_col].apply(lambda x: '已清算' if x else '')

        # Step 2: Replace with '低消耗' if column 11 < 30 and current status is '正常'
        col11 = pd.to_numeric(group[group.columns[11]],errors='coerce').fillna(0)
        low_cost_mask = (
                (col11 < 30) &
                (group[status_col] == '正常')  # Only override if still '正常'
        )
        group.loc[low_cost_mask, status_col] = '低消耗'
        # Step 3: Replace with '需充值' if column 12 < 0.1 * column 13 and status is not '异常'


        # Use DataFrame expressions directly to preserve row alignment
        mask = (
                       pd.to_numeric(group[remaining_col], errors='coerce').fillna(0).astype(int) < 0
               ) & (group[status_col] == '正常')

        # Apply update
        group.loc[mask, status_col] = '需充值'

        sum_indices = [1, 2, 4, 6, 8, 11, 12, 13, 14, 15]  # 查看人数, 查看数, 点击数, 注册数, 充值数 , 花销, 剩余额度 , 已充额度
        avg_indices = [3, 5, 7, 9]     # 人均查看, 单次点击, 单次注册， 单次充值，

        # Create summary row
        summary_row = [''] * len(group.columns)
        summary_row[0] = '合计'

        for i in sum_indices:
            col = group.columns[i]
            # Convert values to numbers, coerce errors to NaN, then drop them
            numeric_values = pd.to_numeric(group[col], errors='coerce').dropna()
            if not numeric_values.empty:
                summary_row[i] = round(numeric_values.sum(), 2)

        for i in avg_indices:
            col = group.columns[i]
            numeric_values = pd.to_numeric(group[col], errors='coerce')
            valid_values = numeric_values.dropna()

            if not valid_values.empty:
                summary_row[i] = round(valid_values.sum() / len(valid_values), 2)

        # Ensure summary_row has the same number of columns as group (including 'Deposit')
        if len(summary_row) < len(group.columns):
            summary_row.append('')  # Add empty value for the new 'Deposit' column

        group.loc[len(group)] = summary_row
        group.loc[len(group)] = [''] * len(group.columns)
        group.loc[len(group)] = [''] * len(group.columns)

        grouped_rows.append(group)

    final_df = pd.concat(grouped_rows, ignore_index=True)
    final_df.drop(final_df.columns[18],axis=1,inplace=True)

    # ---- Final Summary Row for All ----
    total_summary = [''] * len(final_df.columns)
    total_summary[0] = '总计'

    # same indices used earlier
    sum_indices = [1, 2, 4, 6, 8, 11, 12, 13, 14, 15]
    avg_indices = [3, 5, 7, 9]

    for i in sum_indices:
        col = df.columns[i]
        numeric_values = pd.to_numeric(df[col], errors='coerce').dropna()
        if not numeric_values.empty:
            total_summary[i] = round(numeric_values.sum(), 2)

    for i in avg_indices:
        col = df.columns[i]
        numeric_values = pd.to_numeric(df[col], errors='coerce').dropna()
        if not numeric_values.empty:
            total_summary[i] = round(numeric_values.mean(), 2)

    # Append final summary and spacer row
    final_df.loc[len(final_df)] = total_summary


    final_df['剩余储备金'] = ''
    value = final_df.iloc[-1, 15]
    final_df.at[0, '剩余储备金'] = f'{round(-value + bm_ins["deposit"],2)}'
    final_df.at[1, '剩余储备金'] = '更新时间'
    # get now time  MM-DD HH:SS
    now = datetime.now()
    final_df.at[2, '剩余储备金'] = now.strftime('%m-%d %H:%M')

    final_df['报告类型'] = ''
    final_df.at[0, '报告类型'] = '总消耗'


    # Output paths
    xlsx_path = os.path.join(output_folder, f"{bm_ins['business_name']}.xlsx")

    # Remove old files
    for path in [xlsx_path]:
        if os.path.exists(path):
            os.remove(path)

    # Save Excel
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=bm_ins['business_name'], index=False)

        # mark col12 (index 12) with red background if value is negative
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl import load_workbook

        workbook = writer.book
        worksheet = writer.sheets[bm_ins['business_name']]

        # 🔹 set first column width
        worksheet.column_dimensions['A'].width = 32

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        gray_fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
        blue_fill  = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")
        purple_fill = PatternFill(start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid")

        # Apply background color based on column 16 value
        for row_idx in range(len(final_df)):
            status = final_df.iloc[row_idx, 16]
            fill = None
            if status == '低消耗':
                fill = gray_fill
            elif status == '需充值':
                fill = blue_fill
            elif status == '异常':
                fill = red_fill

            if fill:

                excel_row = row_idx + 2  # Excel data starts from row 2
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}{excel_row}"]
                    cell.fill = fill

        for row_idx in range(len(final_df)):
            cleared = final_df.iloc[row_idx, 17]
            if cleared == '已清算':
                fill = purple_fill

                excel_row = row_idx + 2  # Excel data starts from row 2
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}{excel_row}"]
                    cell.fill = fill

        print(bm_ins['business_name'], ' done')


def writeYesterdayReport(report, bm_ins,sheet_name='Report'):

    concated_report = yesterdaySummaryTable(report,bm_ins['business_name'])
    headers = summary_headers + yesterday_concat_headers
    try:
        df = pd.DataFrame(data=concated_report, columns=headers)
    except:
        df = None

    df['display'] = ''

    status_col = df.columns[12]
    cleared_col = df.columns[13]
    display_col = df.columns[14]

    df[display_col] = df[[status_col,cleared_col]].apply(lambda x: not (x[0] == True and x[1] == False), axis=1)

    grouped_rows = []

    for display, group in df.groupby(display_col):
        group = group.copy().reset_index(drop=True)

        # Convert values that can be numeric
        for col in group.columns:
            group[col] = pd.to_numeric(group[col], errors='ignore')


        status_col = group.columns[12]
        cleared_col = group.columns[13]

        # Step 1: Mark '正常' or '异常' based on original column 14 values
        group[status_col] = group[status_col].apply(lambda x: '正常' if x else '异常')
        group[cleared_col] = group[cleared_col].apply(lambda x: '已清算' if x else '')

        # Step 2: Replace with '低消耗' if column 11 < 30 and current status is '正常'
        col11 = pd.to_numeric(group[group.columns[11]],errors='coerce').fillna(0)
        low_cost_mask = (
                (col11 < 30) &
                (group[status_col] == '正常')  # Only override if still '正常'
        )
        group.loc[low_cost_mask, status_col] = '低消耗'
        # Step 3: Replace with '需充值' if column 12 < 0.1 * column 13 and status is not '异常'


        sum_indices = [1, 2, 4, 6, 8, 11,]  # 查看人数, 查看数, 点击数, 注册数, 充值数 , 花销,
        avg_indices = [3, 5, 7, 9]     # 人均查看, 单次点击, 单次注册， 单次充值，

        # Create summary row
        summary_row = [''] * len(group.columns)
        summary_row[0] = '合计'

        for i in sum_indices:
            col = group.columns[i]
            # Convert values to numbers, coerce errors to NaN, then drop them
            numeric_values = pd.to_numeric(group[col], errors='coerce').dropna()
            if not numeric_values.empty:
                summary_row[i] = round(numeric_values.sum(), 2)

        for i in avg_indices:
            col = group.columns[i]
            numeric_values = pd.to_numeric(group[col], errors='coerce')
            valid_values = numeric_values.dropna()

            if not valid_values.empty:
                summary_row[i] = round(valid_values.sum() / len(valid_values), 2)

        # Ensure summary_row has the same number of columns as group (including 'Deposit')
        if len(summary_row) < len(group.columns):
            summary_row.append('')  # Add empty value for the new 'Deposit' column

        group.loc[len(group)] = summary_row
        group.loc[len(group)] = [''] * len(group.columns)
        group.loc[len(group)] = [''] * len(group.columns)

        grouped_rows.append(group)

    final_df = pd.concat(grouped_rows, ignore_index=True)
    final_df.drop(final_df.columns[14],axis=1,inplace=True)

    # ---- Final Summary Row for All ----
    total_summary = [''] * len(final_df.columns)
    total_summary[0] = '总计'

    # same indices used earlier
    sum_indices = [1, 2, 4, 6, 8, 11,]
    avg_indices = [3, 5, 7, 9]

    for i in sum_indices:
        col = df.columns[i]
        numeric_values = pd.to_numeric(df[col], errors='coerce').dropna()
        if not numeric_values.empty:
            total_summary[i] = round(numeric_values.sum(), 2)

    for i in avg_indices:
        col = df.columns[i]
        numeric_values = pd.to_numeric(df[col], errors='coerce').dropna()
        if not numeric_values.empty:
            total_summary[i] = round(numeric_values.mean(), 2)

    # Append final summary and spacer row
    final_df.loc[len(final_df)] = total_summary

    # remove column 7
    final_df.loc[len(final_df)] = [''] * len(final_df.columns)

    final_df['报告类型'] = ''
    final_df.at[0, '报告类型'] = '昨日消耗'


    # Output paths
    xlsx_path = os.path.join(output_folder, f"Y_{sheet_name}.xlsx")

    # Remove old files
    for path in [xlsx_path]:
        if os.path.exists(path):
            os.remove(path)

    # Save Excel
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # mark col12 (index 12) with red background if value is negative
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl import load_workbook

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 🔹 set first column width
        worksheet.column_dimensions['A'].width = 32

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        gray_fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
        blue_fill  = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")
        purple_fill = PatternFill(start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid")

        # Apply background color based on column 16 value
        for row_idx in range(len(final_df)):
            status = final_df.iloc[row_idx, 12]
            fill = None
            if status == '低消耗':
                fill = gray_fill
            elif status == '需充值':
                fill = blue_fill
            elif status == '异常':
                fill = red_fill

            if fill:

                excel_row = row_idx + 2  # Excel data starts from row 2
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}{excel_row}"]
                    cell.fill = fill

        for row_idx in range(len(final_df)):
            cleared = final_df.iloc[row_idx, 13]
            if cleared == '已清算':
                fill = purple_fill

                excel_row = row_idx + 2  # Excel data starts from row 2
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}{excel_row}"]
                    cell.fill = fill

        print(sheet_name, ' done')


def concatSummaryTable(report, bm_name):
    new_report = []
    for item in report:
        account = findAccount(records, item[10],bm_name)
        # print(account,"\n")
        # replace record name with account name
        if account:
            item[0] = account['name']

        status_account = findAccount(clients, item[10],bm_name)
        status = status_account['option'] if status_account else False
        cleared = status_account['cleared'] if status_account else False
        cash = getCashAmount(account) if account else 0
        deposit = getDepositAmount(account) if account else 0

        # Append placeholders and calculated fields
        item.pop(-1) # last item is which read from selenium but empty

        # replace spent with status account spent when it is cleared
        if cleared:
            # spent
            item[-1] = status_account['spent'] if status_account and 'spent' in status_account else item[-1]
        item.extend([
            0,      # remaining
            0,      # total
            cash,
            deposit,
            status,   # status
            cleared,    # cleared
        ])
        new_report.append(item)

    ret = new_report
    # ret = [report for report in new_report if report[-1] == False and report[-2] == True] # 已清算的不计入
    # printElement(ret)
    return ret

def yesterdaySummaryTable(report, bm_name):
    new_report = []
    for item in report:
        status_account = findAccount(clients, item[10],bm_name)
        status = status_account['option'] if status_account else False
        cleared = status_account['cleared'] if status_account else False

        # Append placeholders and calculated fields
        item.pop(-1) # last item is which read from selenium but empty
        item.extend([
            status,   # status
            cleared,    # cleared
        ])
        new_report.append(item)

    ret = new_report
    # ret = [report for report in new_report if report[-1] == False and report[-2] == True] # 已清算的不计入
    # printElement(ret)

    return ret


def findAccount(accounts,account_id,bm_name):
    for account in accounts:
        if account['id'] == account_id and account['name'].startswith(bm_name):
            return account
    return None

def matchRecord(records,item):
    for record in records:
        if item['id'] == record['id'] and item['date'] == record['date']:
            return item
    return None

def getCashAmount(account):
    amount = 0
    for item in account['records']:
        if not item['deposit']: amount += item['amount']
    return amount

def getDepositAmount(account):
    amount = 0
    for item in account['records']:
        if item['deposit']: amount += item['amount']
    return amount


def flatChargeRecords():
    flat = []
    for item in records:
        new_item = deepcopy(item)
        del new_item['records']

        for charge_record in item['records']:
            new_record = deepcopy(new_item)
            date = convert_to_full_date(charge_record['date'])
            amount = charge_record['amount']
            deposit = charge_record['deposit']
            new_record['date'] = date
            new_record['amount'] = amount
            new_record['deposit'] = deposit
            flat.append(new_record)

    return flat

def getTotalChargeAmount(flat_records):
    amounts = {}
    for item in flat_records:
        account_id = item['id']
        if account_id not in amounts:
            amounts[account_id] = 0
        amounts[account_id] += item['amount']
    return amounts

def appendChargeAmount(report,flat_records):
    """ append charge amount to report """
    amounts = getTotalChargeAmount(flat_records)
    for item in report:
        account_id = item[10]
        if account_id in amounts:
            item.append(amounts[account_id])
        else:
            item.append(0)  # if no charge amount, append 0
    return report



def concatEverydaySummaryTable(report,bm_name):
    new_report = []
    for item in report:
        account = findAccount(records, item[10], bm_name)
        status_account = findAccount(clients, item[10],bm_name)
        status = status_account['option'] if status_account else False
        cleared = status_account['cleared'] if status_account else False
        cash = getCashAmount(account) if account else 0
        deposit = getDepositAmount(account) if account else 0

        # Append placeholders and calculated fields
        item.pop(-1) # last item is which read from selenium but empty
        item.extend([
            0,      # remaining
            0,      # total
            cash,
            deposit,
            status,   # status
            cleared    # cleared
        ])
        new_report.append(item)

    return new_report

def calcTotalDeposit(report,flat_records):
    """ list out all id in report and find in charge_map, then sum up the deposit amount """
    total_deposit = 0
    ids = set()
    for item in report:
        ids.add(item[10])
    for id in ids:
        for record in flat_records:
            if record['id'] == id and record['deposit']:
                total_deposit += record['amount']

    return total_deposit


def writeEverydayReport(report, bm_ins,sheet_name='Report'):
    flat_records = flatChargeRecords()
    charge_map = collectChargesByAccount(flat_records)
    status_map = collectStatusByAccount()

    total_deposit = calcTotalDeposit(report,flat_records)

    report = appendChargeAmount(report, flat_records)


    headers = everyday_headers + everyday_concat_headers
    # printElement(report)

    checkRow(report,correct_element_number=15)

    df = pd.DataFrame(data=report, columns=headers)


    # Add extra charge columns
    df['充值'] = ''
    df['充值日期'] = ''
    df['储备金'] = ''

    # Ensure numeric types
    col14 = pd.to_numeric(df[df.columns[14]], errors='coerce').fillna(0)  # charged
    col11 = pd.to_numeric(df[df.columns[11]], errors='coerce').fillna(0)  # spent
    df[df.columns[12]] = ((col14 - col11) * 100).astype(int) / 100
    df[df.columns[12]] = df[df.columns[12]].apply(lambda x: '-' if x == 0 else round(x, 2))

    account_id_col = df.columns[10]
    grouped_rows = []

    for account_id, group in df.groupby(account_id_col):
        group = group.copy().reset_index(drop=True)

        # Calculate remaining deposit
        remaining_deposit = []
        initial_charge = pd.to_numeric(group.iloc[0][df.columns[14]], errors='coerce')
        spent_series = pd.to_numeric(group[df.columns[11]], errors='coerce').fillna(0)

        remaining = initial_charge - spent_series.iloc[0]
        if abs(remaining) < 0.01:
            remaining = 0
        remaining_deposit.append(remaining)
        for i in range(1, len(group)):
            remaining -= spent_series.iloc[i]
            remaining_deposit.append(remaining)

        group[df.columns[12]] = [round(x, 2) for x in remaining_deposit]

        # Hide duplicate charged value
        for i in range(1, len(group)):
            group.at[i, df.columns[14]] = ''

        # Inject charge strings into first row
        charges = charge_map.get(account_id, [])
        if charges:
            amount_lines = '\n'.join(str(c[0]) for c in charges)
            date_lines = '\n'.join(str(c[1]) for c in charges)
            deposit_lines = '\n'.join(str(c[2]) for c in charges)

            group.at[0, '充值'] = amount_lines
            group.at[0, '充值日期'] = date_lines
            group.at[0, '储备金'] = deposit_lines

        # Summary rows
        sum_indices = [1, 2, 4, 6, 8, 11]
        avg_indices = [3, 5, 7, 9]
        group_summary_row = [''] * (len(group.columns))
        group_summary_row[0] = f'合计-{account_id}'

        for i in sum_indices:
            col = group.columns[i]
            numeric_values = pd.to_numeric(group[col], errors='coerce').dropna()
            if not numeric_values.empty:
                group_summary_row[i] = round(numeric_values.sum(), 2)

        for i in avg_indices:
            col = group.columns[i]
            numeric_values = pd.to_numeric(group[col], errors='coerce').dropna()
            if not numeric_values.empty:
                group_summary_row[i] = round(numeric_values.mean(), 2)

        group_summary_row[12] = '' if status_map[account_id]['option'] else '异常'
        group_summary_row[13] = '已结算' if status_map[account_id]['cleared'] else ''

        group.loc[len(group)] = group_summary_row
        group.loc[len(group)] = [''] * len(group.columns)


        grouped_rows.append(group)

    final_df = pd.concat(grouped_rows, ignore_index=True)

    final_df['剩余储备金'] = ''  # Add a column for Deposit, initialized with empty strings
    # Populate the last row with the "Deposit" title and the deposit value
    final_df.at[0, '剩余储备金'] = f'{round(-total_deposit + bm_ins["deposit"],2)}'  # Add Deposit title in the last row of "Deposit" column
    final_df.at[1, '剩余储备金'] = '更新时间'
    # get now time  MM-DD HH:SS
    now = datetime.now()
    final_df.at[2, '剩余储备金'] = now.strftime('%m-%d %H:%M')
    # final_df.at[3, '剩余储备金'] = '报表截止'
    # final_df.at[4, '剩余储备金'] = get_end_date().strftime('%m-%d')

    # Output patars
    xlsx_path = os.path.join(output_folder, f"{sheet_name}.xlsx")

    for path in [xlsx_path]:
        if os.path.exists(path):
            os.remove(path)

    # Save Excel with openpyxl
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        purple_fill = PatternFill(start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid")

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 🔹 set first column width
        worksheet.column_dimensions['A'].width = 32

        # Loop through cells to apply formatting
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.value == "异常":
                    cell.fill = red_fill
                elif cell.value == "已结算":
                    cell.fill = purple_fill


    # Wrap charge cells
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    wrap_columns = ['充值', '充值日期', '储备金']
    col_indices = [i + 1 for i, col in enumerate(final_df.columns) if col in wrap_columns]

    for row in ws.iter_rows(min_row=2):  # skip header
        for idx in col_indices:
            cell = row[idx - 1]
            cell.alignment = Alignment(wrap_text=True)

    for idx in col_indices:
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = 15

    wb.save(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active


    # Store the cells to split first to avoid mutation while iterating
    cells_to_split = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and '\n' in cell.value:
                cells_to_split.append((cell.coordinate, cell.value.strip().split('\n')))

    # Now split each collected cell downward
    for coord, lines in cells_to_split:
        col = ''.join(filter(str.isalpha, coord))  # Extract column letter(s)
        row = int(''.join(filter(str.isdigit, coord)))  # Extract row number
        for i, line in enumerate(lines):
            ws[f"{col}{row + i}"].value = line.strip()

    # Save final result
    wb.save(xlsx_path)

    print(sheet_name, ' done')



def safe_div(numerator, denominator, digits=2):
    try:
        if not denominator:
            return 0
        getcontext().prec = digits + 5  # set precision higher than needed
        result = Decimal(str(numerator)) / Decimal(str(denominator))
        return float(result.quantize(Decimal('1.' + '0' * digits), rounding=ROUND_DOWN))
    except:
        return 0

def safe_float(v):
    try:
        return float(v)
    except:
        return 0

def writeEverydayAdsetReport(report, sheet_name='Report'):

    # Validate rows
    checkRow(report, correct_element_number=11)

    # Create DataFrame
    df = pd.DataFrame(data=report, columns=adset_headers)

    # Check column index exists
    if len(df.columns) <= 11:
        raise IndexError("Expected at least 12 columns, but got fewer.")

    adset_id_col = df.columns[11]
    grouped_rows = []

    for adset_id, group in df.groupby(adset_id_col):
        group = group.copy().reset_index(drop=True)

        sum_indices = [1, 2, 4, 6, 8, 10]  # Columns to be summed
        group_summary_row = [''] * len(group.columns)
        group_summary_row[0] = f'合计-{adset_id}'

        # Compute sums
        for i in sum_indices:
            col = group.columns[i]
            values = pd.to_numeric(group[col], errors='coerce')
            group_summary_row[i] = round(values.sum(), 2)


        group_summary_row[3] = safe_div(group_summary_row[2], group_summary_row[1],4) * 100
        group_summary_row[5] = safe_div(group_summary_row[10], group_summary_row[4])
        group_summary_row[7] = safe_div(group_summary_row[10], group_summary_row[6])
        group_summary_row[9] = safe_div(group_summary_row[10], group_summary_row[8])

        # Add to group
        group.loc[len(group)] = group_summary_row
        group.loc[len(group)] = [''] * len(group.columns)  # empty row after summary

        grouped_rows.append(group)

    # Concatenate all groups
    final_df = pd.concat(grouped_rows, ignore_index=True)
    final_df.iloc[:,11] = final_df.iloc[:,13]
    final_df.drop(columns=['占位1', '占位2'], inplace=True, errors='ignore')


    # Output paths
    xlsx_path = os.path.join(output_folder, f"{sheet_name}.xlsx")


    # Save Excel with openpyxl
    # Save Excel with openpyxl
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        green_fill = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 🔹 set first column width
        worksheet.column_dimensions['A'].width = 32

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            first_cell = row[0].value
            if isinstance(first_cell, str) and first_cell.startswith('合计'):
                try:
                    purchase = safe_float(row[7].value)
                    trial_started = safe_float(row[9].value)
                    spent = safe_float(row[10].value)
                except (ValueError, TypeError):
                    continue

                purchase_kpi_check = KPIlevel(purchase, spent)
                print(purchase_kpi_check)
                if purchase_kpi_check == 'red':
                    row[7].fill = red_fill
                elif purchase_kpi_check == 'yellow':
                    row[7].fill = yellow_fill
                elif purchase_kpi_check == 'green':
                    row[7].fill = green_fill

                trial_kpi_check = KPIlevel(trial_started, spent)
                if trial_kpi_check == 'red':
                    row[9].fill = red_fill
                elif trial_kpi_check == 'yellow':
                    row[9].fill = yellow_fill
                elif trial_kpi_check == 'green':
                    row[9].fill = green_fill



    print(sheet_name, 'done')

def KPIlevel(money,spent):
    if spent < 15:
        return 'blank'
    else:
        if money == 0:
            return 'red'
        elif money < 15:
            return 'green'
        elif money < 30:
            return 'yellow'
        else:
            return 'red'



def writeAccountReport(report,sheet_name='accounts'):

    df = pd.DataFrame(data=report)

    df = df[["name","spending_limit","threshold","status","currency",]]
    df["status"] = df["status"].replace(False,"异常")
    df["status"] = df["status"].replace(True,"")

    df = df.rename(columns=account_info_headers)

    xlsx_path = os.path.join(output_folder, f"{sheet_name}.xlsx")

    for path in [xlsx_path]:
        if os.path.exists(path):
            os.remove(path)

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def checkRow(reports, correct_element_number):
    for i,row in enumerate(reports,1):
        if len(row) != correct_element_number:
            print(i,row)

def yesterdayFinanceReport(yepay_data):
    # 日期 | 充值金额 / 笔数 | 服务费 | 取消授权费 / 笔数 | 完成金额 /笔数 | 授权金额/占比 | 撤销金额/占比 | 拒付金额/占比 | 授权成功率/笔数 | 撤销笔数/撤销率 | 拒付笔数/拒付率
    #  0            1         2             3              4              5            6               7             8             9                10
    yepay_data = ['2025-09-05', '33501.53/5', '167.5', '8.4/28', '18382.46/142', '20725.49/74.82%', '679/2.45%', '6294.8/22.73%',
     '149/84.18%', '2/1.13%', '26/14.69%']

    # 日期 充值金额 转账费1u  卡台千五  取消费  取消笔数 撤销笔数 拒付笔数


def totalCharge():
    # count the total charge amount in records, which is deposit =False
    total = 0
    for account in records:
        for record in account['records']:
            if not record['deposit']:
                total += record['amount']



def chargeIt(account,report,bm_ins):
    # account is BM ins account
    # report is fb report data
    # record is charge records

    new_report = []

    # loop chargeIt each account is bm_ins account

    for record in records: # walk charges
        # first find charge records
        if account['name'] == record['name'] and account['id'] == record['id']: # find account charges
            # base is charge, fill every charge to each account
            for charge in record['records']:

                filled = False
                idx = 0
                data = {}
                while not filled and idx < len(report):
                    data = report[idx]
                    data['cash'] = 0
                    data['cash_count'] = 0
                    data['deposit'] = 0
                    data['deposit_count'] = 0

                    if data['Account Name'] == account['name'] and data['Account ID'] == account['id'] and charge['date'] == data['date']:

                        if charge['deposit']:
                            data['deposit'] += charge['amount']
                            data['deposit_count'] += 1
                        else:
                            data['cash'] += charge['amount']
                            data['cash_count'] += 1

                        filled = True
                        del report[idx]

                    idx += 1

                if not filled:
                    for field_name in bm_ins['header']:
                        data[field_name] = "-"
                    data["Account Name"] = account['name']
                    data["Account ID"] = account['id']

                    data['cash'] = 0
                    data['cash_count'] = 0
                    data['deposit'] = 0
                    data['deposit_count'] = 0

                    if charge['deposit']:
                        data['deposit'] += charge['amount']
                        data['deposit_count'] += 1
                    else:
                        data['cash'] += charge['amount']
                        data['cash_count'] += 1

                    data['date'] = charge['date']


                new_report.append(data)
    return new_report

def getCharge(account):
    charges = []

    for record in records:
        if account['name'] != record['name'] or account['id'] != record['id']:
            continue

        last_charge = None

        for charge in record['records']:
            # New date → close previous record
            if last_charge is None or charge['date'] != last_charge['date']:
                if last_charge is not None:
                    charges.append(last_charge)

                last_charge = {
                    "name": record['name'],
                    "id": record['id'],
                    "date": charge['date'],
                    "cash": 0,
                    "cash_count": 0,
                    "deposit": 0,
                    "deposit_count": 0,
                }

            # Accumulate
            if charge['deposit']:
                last_charge['deposit'] += charge['amount']
                last_charge['deposit_count'] += 1
            else:
                last_charge['cash'] += charge['amount']
                last_charge['cash_count'] += 1

        # Append final day
        if last_charge is not None:
            charges.append(last_charge)

    return charges

def StdDate(date):
    try:
        dt = datetime.strptime(date, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return date

def concatCharges(charges, report, bm_ins ,account):
    try:
        if account['spent']:
            new_item = {field: "-" for field in bm_ins['header']}
            new_item['Amount spent'] = account['spent']
            new_item["Account Name"] = account['name']
            new_item["Account ID"] = account['id']
            report.insert(0,new_item)
    except:
        pass

    for item in charges:
        item['date'] = StdDate(item['date'])
    for item in report:
        item['date'] = StdDate(item['date'])

    new_report = []

    matched_charges = []
    matched_report = []

    for charge in charges:

        # 如果report 里面找到了 符合条件的记录 把这个记录添加charge
        for item in report:
            print(item)
            if (
                item['date'] == charge['date']
                and item['Account Name'] == account['name']
                and charge['name'] == account['name']
            ):
                new_item = item.copy()
                new_item['cash'] = charge['cash']
                new_item['cash_count'] = charge['cash_count']
                new_item['deposit'] = charge['deposit']
                new_item['deposit_count'] = charge['deposit_count']


                new_report.append(new_item)
                matched_charges.append(new_item)
                matched_report.append(new_item)
                break


    report_map = {
        (item['date'], item['Account Name']): item
        for item in matched_charges
    }


    for charge in charges:
        key = (charge['date'], charge['name'])
        if charge['name'] == account['name'] and key not in report_map:

            new_item = {field: "-" for field in bm_ins['header']}
            new_item["Account Name"] = account['name']
            new_item["Account ID"] = account['id']
            new_item["date"] = charge['date']
            new_item['cash'] = charge['cash']
            new_item['cash_count'] = charge['cash_count']
            new_item['deposit'] = charge['deposit']
            new_item['deposit_count'] = charge['deposit_count']

            report_map[key] = new_item.copy()

    for item in report:
        key = (item['date'], item['Account Name'])
        if item['Account Name'] == account['name'] and key not in report_map:
            item['cash'] = 0
            item['cash_count'] = 0
            item['deposit'] = 0
            item['deposit_count'] = 0

            report_map[key] = item.copy()

    for item in report_map.values():
        key = (item['date'], item['Account Name'])
        report_map[key]['time'] = datetime.now()
        report_map[key]['hash'] = genHash(report_map[key])

    return list(report_map.values())


def genHash(item):
    hash = item['date'] + item['Account Name']
    return hash

def writeIt(report, bm_ins):
    all_report = []

    all_hash = {}

    for account in bm_ins['accounts']:
        charges = getCharge(account)
        new_report = concatCharges(charges,report,bm_ins,account)
        for item in new_report:
            if item['hash'] not in all_hash.keys():
                all_hash[item['hash']] =  item['time']
                all_report.append(item)
            else:
                if item['time'] > all_hash[item['hash']]:
                    all_hash[item['hash']] = item['time']
                    # update all_report same hash item
                    for idx, old_item in enumerate(all_report):
                        if old_item['hash'] == item['hash']:
                            all_report[idx] = item
                            break

    all = sorted(all_report, key=lambda x: x['Account Name'])

    writeDayToDayReport(all,bm_ins)

def updateIt(report, bm_ins):
    all_report = []
    all_hash = {}

    data = retrieve_previous_data(bm_ins)
    report.extend(data)

    print(report)

    for account in bm_ins['accounts']:
        charges = getCharge(account)
        new_report = concatCharges(charges,report,bm_ins,account)
        for item in new_report:
            if item['hash'] not in all_hash.keys():
                all_hash[item['hash']] =  item['time']
                all_report.append(item)
            else:
                if item['time'] > all_hash[item['hash']]:
                    all_hash[item['hash']] = item['time']
                    # update all_report same hash item
                    for idx, old_item in enumerate(all_report):
                        if old_item['hash'] == item['hash']:
                            all_report[idx] = item
                            break

    all = sorted(all_report, key=lambda x: x['Account Name'])

    writeDayToDayReport(all,bm_ins)


def writeDayToDayReport(data,bm_ins):
    df = pd.DataFrame(data)

    df['_date_plain'] = df['date']
    df['_date'] = df['date'].apply(normalize_date)
    df['_amount_spent'] = df['Amount spent'].apply(normalize_amount)

    summary_all = []
    summary_before = []
    detail_rows = []

    # ---------- PER ACCOUNT ----------
    for account_name, group in df.groupby('Account Name'):
        group = group.sort_values('_date', ascending=False)

        before_today = group[(group['_date'] < TODAY) | (group['_date'].isna())]
        total_summary = make_summary(group, account_name, "最新")
        before_summary = make_summary(before_today, account_name,"截止昨日")

        total_summary['Account ID'] = group['Account ID'].iloc[0]
        before_summary['Account ID'] = group['Account ID'].iloc[0]

        summary_all.append(total_summary)
        summary_before.append(before_summary)

        # detail section
        detail_rows.append(make_detail_header())

        detail_rows.append(total_summary)
        detail_rows.append(before_summary)

        detail_rows.extend(
            group.drop(columns=['_date', '_amount_spent']).to_dict('records')
        )
        detail_rows.append({})

    # ---------- SUMMARY Everyday ----------
    date_summary_all = []

    for _date_plain, group in df.groupby('_date_plain'):
        group = group.sort_values('Account Name', ascending=False)
        date_summary = make_date_summary(group,_date_plain, bm_ins['business_name'])
        date_summary_all.append(date_summary)


    date_summary_df = pd.DataFrame(date_summary_all)
    date_summary_df = date_summary_df.rename(columns=make_date_header())


    # ---------- SUMMARY TABLES ----------

    summary_all_df = pd.DataFrame(summary_all)
    summary_before_df = pd.DataFrame(summary_before)

    grand_all = make_summary_from_summary(
        summary_all_df, "最新消耗"
    )

    grand_before = make_summary_from_summary(
        summary_before_df, "截至昨日"
    )

    summary_all_df = pd.concat(
        [pd.DataFrame([grand_all]), summary_all_df],
        ignore_index=True
    )

    summary_all_df['剩余储备金'] = ''  # Add a column for Deposit, initialized with empty strings
    # Populate the last row with the "Deposit" title and the deposit value
    summary_all_df.at[0, '剩余储备金'] = f'{round(-grand_all["deposit"] + bm_ins["deposit"],2)}'  # Add Deposit title in the last row of "Deposit" column
    summary_all_df.at[1, '剩余储备金'] = '更新时间'
    # get now time  MM-DD HH:SS
    now = datetime.now()
    summary_all_df.at[2, '剩余储备金'] = now.strftime('%m-%d %H:%M')

    summary_before_df = pd.concat(
        [pd.DataFrame([grand_before]), summary_before_df],
        ignore_index=True
    )

    detail_df = pd.DataFrame(detail_rows)

    # ---------- RENAME HEADER ----------

    summary_all_df = summary_all_df.rename(columns=HEADER_MAP)


    summary_before_df = summary_before_df.rename(columns=HEADER_MAP)
    detail_df = detail_df.rename(columns=HEADER_MAP)

    # summary_all_df.drop(columns=['time'], inplace=True)
    # summary_before_df.drop(columns=['time'], inplace=True)
    detail_df.drop(columns=['time'], inplace=True)
    detail_df.drop(columns=['hash'], inplace=True)
    detail_df.drop(columns=['_date_plain'], inplace=True)

    # ---------- Drop time -----------



    # ---------- WRITE ONE SHEET ----------

    output_folder = "statistics"
    filename = f"{bm_ins['business_name']}.xlsx"
    output_path = os.path.join(output_folder, filename)


    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        start_row = 0

        summary_all_df.to_excel(
            writer, sheet_name="REPORT", startrow=start_row, index=False
        )

        start_row += len(summary_all_df) + 2

        summary_before_df.to_excel(
            writer, sheet_name="REPORT", startrow=start_row, index=False
        )

        start_row += len(summary_before_df) + 3

        detail_df.to_excel(
            writer, sheet_name="REPORT", startrow=start_row, index=False
        )

        start_row += len(detail_df) + 3

        date_summary_df.to_excel(
            writer, sheet_name="REPORT", startrow=start_row, index=False
        )

    wb = load_workbook(output_path)
    ws = wb["REPORT"]

    auto_adjust_column_width(ws)

    bold = Font(bold=True)
    purple_bold_font = Font(color="C05780", bold=True)
    blue_bold_font = Font(color="0065A2", bold=True)

    # --- find column index of "date" dynamically ---
    header_row = [cell.value for cell in ws[1]]

    date_col_idx = header_row.index("日期")  # or "date" if not renamed yet

    # --- apply formatting ---
    for row in ws.iter_rows(min_row=2):
        cell_value = row[date_col_idx].value

        if cell_value and "所有账户" in str(cell_value):
            for cell in row:
                cell.font = bold

        elif cell_value and "最新" in str(cell_value):
            for cell in row:
                cell.font = purple_bold_font

        elif cell_value and "截止昨日" in str(cell_value):
            for cell in row:
                cell.font = blue_bold_font

    wb.save(output_path)


def auto_adjust_column_width(ws, min_width=12):
    header_row = ws[1]  # 第一行是表头

    for col_idx, cell in enumerate(header_row, start=1):
        col_letter = get_column_letter(col_idx)

        width = min_width

        if cell.value:
            try:
                width = WIDTH_MAP[cell.value]
            except Exception:
                width = min_width

        ws.column_dimensions[col_letter].width =  width



HEADER_MAP = {
    "Account Name": "广告账户名称",
    "Account ID": "账户ID",
    "Amount spent": "已用额度",
    "remaining": "剩余额度",
    "total": "总充值额度",
    "date": "日期",
    "cash": "现金充值",
    "cash_count": "次数",
    "deposit": "备用金充值",
    "deposit_count": "次数",
}

WIDTH_MAP = {
    "广告账户名称": 32,
    "账户ID": 20,
    "已用额度": 12,
    "剩余额度": 12,
    "总充值额度": 12,
    "日期": 11,
    "现金充值": 12,
    "次数": 11,
    "备用金充值": 12,
    "剩余储备金": 12,
}

TODAY = pd.to_datetime(date.today())

def normalize_date(d):
    if d == '-' or pd.isna(d):
        return pd.NaT
    return pd.to_datetime(d)

def normalize_amount(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return 0.0

def make_detail_header():
    return {
        "Account Name": "广告账户名称",
        "Account ID": "账户ID",
        "date": "日期",
        "Amount spent": "已用额度",
        "cash": "现金充值",
        "cash_count": "次数",
        "deposit": "备用金充值",
        "deposit_count": "次数",
        "total": "总充值额度",
        "remaining": "剩余额度",
    }

def make_summary(group, label, date_label):
    cash = round(group['cash'].sum(), 2)
    deposit = round(group['deposit'].sum(), 2)
    amount_spent = round(group['_amount_spent'].sum(), 2)

    total = round(cash + deposit, 2)
    remaining = round(total - amount_spent, 2)

    return {
        'Account Name': label,
        'Account ID': '',
        'date': date_label,
        'Amount spent': amount_spent,
        'cash': cash,
        'cash_count': group['cash_count'].sum(),
        'deposit': deposit,
        'deposit_count': group['deposit_count'].sum(),
        'total': total,
        'remaining': remaining,
    }


def make_date_summary(group, label, business_name):
    cash = round(group['cash'].sum(), 2)
    deposit = round(group['deposit'].sum(), 2)
    amount_spent = round(group['_amount_spent'].sum(), 2)

    total = round(cash + deposit, 2)

    return {
        'business_name': business_name,
        'Amount spent': amount_spent,
        'Date': label,
        'cash': cash,
        'cash_count': group['cash_count'].sum(),
        'deposit': deposit,
        'deposit_count': group['deposit_count'].sum(),
        'total': total,
    }


def make_date_header():
    return {
        'business_name': "公司",
        "Amount spent": "已用额度",
        "Date": "日期",
        "cash": "现金充值",
        "cash_count": "次数",
        "deposit": "备用金充值",
        "deposit_count": "次数",
        "total": "总充值额度",
    }

def make_summary_from_summary(group, label):
    cash = round(group['cash'].sum(), 2)
    deposit = round(group['deposit'].sum(), 2)
    amount_spent = round(group['Amount spent'].sum(), 2)

    total = round(cash + deposit, 2)
    remaining = round(total - amount_spent, 2)

    return {
        'Account Name': label,
        'Account ID': '',
        'date': '所有账户',
        'Amount spent': amount_spent,
        'cash': cash,
        'cash_count': group['cash_count'].sum(),
        'deposit': deposit,
        'deposit_count': group['deposit_count'].sum(),
        'total': total,
        'remaining': remaining,
    }




import os
import pandas as pd
from datetime import datetime, date

# 请确保你的环境中已定义 genHash 函数，此处保留原逻辑
def genHash(data):
    # 你的哈希生成逻辑
    return hash(str(data))

def retrieve_previous_data(bm_ins):
    """
    Read Excel file {business_name}.xlsx and return rows where
    'date' is not '最新' or '截止昨日'.
    Columns are mapped as:
    'Account Name': col1, 'Account ID': col2, 'date': col3,
    'Amount spent': col4, 'cash': col5, 'cash_count': col6,
    'deposit': col7, 'deposit_count': col8
    If the file does not exist / read fails / no valid data, return empty list.
    """
    output_folder = "statistics"
    filename = f"{bm_ins['business_name']}.xlsx"
    file_path = os.path.join(output_folder, filename)

    # 1. 核心修改：文件不存在直接返回空数据
    if not os.path.exists(file_path):
        return []

    # 2. 异常捕获：文件读取失败（损坏/格式错误）也返回空数据
    try:
        df = pd.read_excel(file_path, header=None)
    except Exception as e:
        print(f"文件读取失败: {e}")
        return []

    # 去除全空行
    df = df.dropna(how='all')

    # 无有效数据时返回空
    if df.empty:
        return []

    # 映射列名（与原逻辑完全一致）
    df.columns = ['Account Name', 'Account ID', 'date',
                  'Amount spent', 'cash', 'cash_count',
                  'deposit', 'deposit_count','total','remaining','bm_deposit']

    # 过滤不符合条件的行
    origin_data = df[~df['date'].isin(['最新', '截止昨日','所有账户','日期'])]

    # 转换为字典列表
    datas = origin_data.to_dict(orient='records')

    # 补充 time 和 hash 字段（原逻辑保留）
    for data in datas:
        data['time'] = datetime.combine(date.today(), datetime.min.time())
        data['hash'] = genHash(data)

    # 返回最终数据（无符合条件数据时自动返回空列表）
    return datas



def yepay_report(data):
    df = pd.DataFrame(data)

    output_folder = "statistics"
    filename = "yepay.xlsx"
    output_path = os.path.join(output_folder, filename)

    os.makedirs(output_folder, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="REPORT",  # ✅ always specify
            index=False,
            startcol=10           # ✅ correct param
        )


def retrieve_business_everyday_spent(bm_ins):
    """
    Read Excel file {business_name}.xlsx and return rows where
    'date' is not '最新' or '截止昨日'.
    Columns are mapped as:
    'Account Name': col1, 'Account ID': col2, 'date': col3,
    'Amount spent': col4, 'cash': col5, 'cash_count': col6,
    'deposit': col7, 'deposit_count': col8
    """
    output_folder = "statistics"
    filename = f"{bm_ins['business_name']}.xlsx"
    file_path = os.path.join(output_folder, filename)
    df = pd.read_excel(file_path, header=None)  # Assuming no header in your Excel

    df = df.dropna(how='all')

    # Map columns to your keys
    df.columns = ['business_name','Amount spent', 'Date', 'cash',
                  'cash_count','deposit', 'deposit_count','total','temp1','temp2','temp3']

    # Filter rows where date is neither '最新' nor '截止昨日'
    origin_data = df[df['business_name'].isin([bm_ins['business_name'],])]

    datas = origin_data.to_dict(orient='records')

    # Convert to list of dicts
    return datas

def write_business_summary_report():
    from informations.fb_bms import rearrange_bms

    all_datas = []

    for bm_ins in rearrange_bms:
        datas = retrieve_business_everyday_spent(bm_ins)
        all_datas.extend(datas)

    business_summary_report(all_datas)


def business_summary_report(data):
    df = pd.DataFrame(data)

    output_folder = "statistics"
    filename = "bm_summary.xlsx"
    output_path = os.path.join(output_folder, filename)

    os.makedirs(output_folder, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="REPORT",  # ✅ always specify
            index=False,
            startcol=0  # ✅ correct param
        )


if __name__ == "__main__":
    write_business_summary_report()



