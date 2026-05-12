from informations.charge_records import records
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def export_full_summary_with_company_totals(records, filename="summary_with_totals.xlsx"):
    """
    - Summarize each account's charge (cash/deposit amounts and counts)
    - Add company prefix (from name before '_')
    - For each company, append total and blank row
    - Export two sheets: 'Detail' and 'Company Summary'
    - Rename headers (Chinese)
    - Format header row bold + centered
    - Set first column width = 25
    """

    # === 1) Build per-name summary ===
    details = []
    for r in records:
        name = r.get("name", "(no_name)")
        cash_amount = deposit_amount = 0.0
        cash_count = deposit_count = 0

        for rec in r.get("records", []):
            amt = rec.get("amount", 0)
            try:
                amt = float(amt)
            except (TypeError, ValueError):
                amt = 0.0

            if rec.get("deposit"):
                deposit_amount += amt
                deposit_count += 1
            else:
                cash_amount += amt
                cash_count += 1

        details.append({
            "name": name,
            "cash_amount": cash_amount,
            "cash_count": cash_count,
            "deposit_amount": deposit_amount,
            "deposit_count": deposit_count
        })

    df = pd.DataFrame(details)

    # === 2) Extract company prefix ===
    df["company"] = df["name"].astype(str).str.split("_").str[0]

    # === 3) Company summary (clean aggregated table) ===
    numeric_cols = ["cash_amount", "cash_count", "deposit_amount", "deposit_count"]
    company_summary = df.groupby("company", as_index=False)[numeric_cols].sum()

    # === 4) Build detailed rows with company totals and blank row ===
    final_rows = []
    companies_in_order = list(dict.fromkeys(df["company"].tolist()))  # preserve order

    for comp in companies_in_order:
        comp_rows = df[df["company"] == comp].copy()

        # add detail rows
        for _, r in comp_rows.iterrows():
            final_rows.append({
                "name": r["name"],
                "cash_amount": r["cash_amount"],
                "cash_count": r["cash_count"],
                "deposit_amount": r["deposit_amount"],
                "deposit_count": r["deposit_count"],
                "company": r["company"],
                "row_type": "detail"
            })

        # add company total row
        totals = comp_rows[numeric_cols].sum()
        final_rows.append({
            "name": f"总计",
            "cash_amount": totals["cash_amount"],
            "cash_count": totals["cash_count"],
            "deposit_amount": totals["deposit_amount"],
            "deposit_count": totals["deposit_count"],
            "company": comp,
            "row_type": "company_total"
        })

        # add blank separator row
        final_rows.append({
            "name": "",
            "cash_amount": "",
            "cash_count": "",
            "deposit_amount": "",
            "deposit_count": "",
            "company": "",
            "row_type": ""
        })

    detail_with_totals = pd.DataFrame(final_rows)

    # === 5) Apply header map (Chinese names) ===
    header_map = {
        "name": "账户名",
        "cash_amount": "现金充值",
        "cash_count": "笔数",
        "deposit_amount": "备用金充值",
        "deposit_count": "笔数"
    }

    # Keep only mapped columns for detail sheet
    detail_export = detail_with_totals[list(header_map.keys())].rename(columns=header_map)

    # For company summary, handle separately
    company_header_map = {
        "company": "公司",
        "cash_amount": "现金充值",
        "cash_count": "笔数",
        "deposit_amount": "备用金充值",
        "deposit_count": "笔数"
    }
    company_export = company_summary.rename(columns=company_header_map)

    # === 6) Write to Excel ===
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        detail_export.to_excel(writer, sheet_name="Detail", index=False)
        company_export.to_excel(writer, sheet_name="Company Summary", index=False)

    # Load workbook to modify formatting
    wb = load_workbook(filename)

    # --- Modify each sheet ---
    for ws in wb.worksheets:
        # Freeze the first row
        ws.freeze_panes = "A2"

        # Set first column width
        ws.column_dimensions['A'].width = 30

        # Optional: center header
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save workbook
    wb.save(filename)


    print(f"✅ Exported and formatted: {filename}")
    return detail_export, company_export


# Example usage
detail_df, company_df = export_full_summary_with_company_totals(records, "summaryCharges.xlsx")
